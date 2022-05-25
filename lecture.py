# ITP Week 3 Day 1 Lecture

# PIP Review

# For Windows, install Python through the Microsoft Store
# For Mac, in your terminal, run these command one by one
# curl https://bootstrap.pypa.io/get-pip.py -o get-pip.py
# python get-pip.py (if that doesn't work python3 get-pip.py )

# Verify it install correctly with
# pip --version

# Excel Spreadsheet

# Not to insult anyone's intelligence, but before we can programmatically work with Excel,
# we need to understand the breakdown/components of Excel in which computers understand them.

# Components

# - Workbook
# - Worksheets
# - Columns (letters) & Rows (numbers)
# - Cell (intersection of row & column)

#------------OPENPYXL-----------

# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).

# -openpyxl is a downloaded module
# -Download & install openpyxl using pip
#   pip install openpyxl

from openpyxl import Workbook

wb = Workbook()
#use the 'type' method on the new variable to verify what kind of data type you are working with
type(wb) # Result-->  <class 'openpyxl.workbook.workbook.Workbook'>

# A workbook is always created with at least one worksheet. You can get it by using the Workbook.active property:

ws = wb.active

# You can create new worksheets using the Workbook.create_sheet() method
# create_sheet takes in a required string for the title and an optional index int

ws1 = wb.create_sheet("Rugrats") # insert at the end (default)
# or
ws2 = wb.create_sheet("Hey Arnold", 0) # insert at first position

# Sheets are given a name automatically when they are created. They are numbered in sequence (Sheet, Sheet1, Sheet2, …). You can change this name at any time with the Worksheet.title property:

ws.title = "New Title"

# Once you gave a worksheet a name, you can get it as a key of the workbook

same_sheet = wb['New Title']

# You can review the names of all worksheets of the workbook with the Workbook.sheetname attribute

wb.sheetnames  # ['Hey Arnold', 'New Title', 'Rugrats']

# Workbook retains a list of worksheets, so we can loop through it!

for sheet in wb:
    print(sheet.title)


# Playing with data

# Accessing one cell - Cells can be accessed directly as keys of the worksheet:

c = ws['A4']

# This will return the cell at A4, or create one if it does not exist yet. Values can be directly assigned:

ws["A4"] = 4 # VSCode has the key programmed..

# There is also the Worksheet.cell() method.

# This provides access to cells using row and column notation:

d = ws.cell(row=4, column=2, value="Whatever I want")

# PRACTICE 

# NOTE: When a worksheet is created in memory, it contains no cells. They are created when first accessed.

# Because of this feature, scrolling through cells instead of accessing them directly will create them all in memory, even if you don’t assign them a value.

for x in range(1, 101):
    for y in range(1, 101):
        ws.cell(row=x, column=y)

# this will create 100x100 cells in memory, for nothing.


# Similarly, we can access just a single row or a single column

colC = ws['C']
row10 = ws[10]

# Columns and rows do not hold value, so if we want to read from them or write to them, we need to loop through them.

for cell in colC:
    colC[cell] = 'new data'

# You can use the Worksheet.iter_rows and/or WOrksheet.iter_cols method to designate a range of columns and rows.

for row in ws.iter_rows(min_row=1, max_row=10, min_col=5, max_col=10): # for every row in between 1-10 row and 5-10 (E-J) column
    for cell in row:
        print(cell)

# SAVING A FILE

# We have been working with a virtualization of a worksheet and so if we don't save it, we lose it!

# NOTE: this is the WORKBOOK not the WORKSHEET

wb.save("something.xlsx")

# AFTERHOURS ROW/COLUMN SLICING