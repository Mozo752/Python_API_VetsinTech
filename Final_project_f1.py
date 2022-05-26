import requests
import json
import openpyxl
from openpyxl import Workbook, load_workbook
import xlsxwriter

# assign url to variable
url = "http://ergast.com/api/f1/2022/last/driverStandings.json"

#check response status of url
response = requests.request("GET", url)

#print response 
print(response)

#print content / response.text
print(response.text)

# #Assign response object to variable (json loads)
clean_data = json.loads(response.text)
# print the cleaned data
# print(clean_data)

#----------Following iterations was used to explore JSON file and how to access keys/lists-------
clean1 = clean_data['MRData']
# print(clean1)

clean2 = clean1['StandingsTable']
# print(clean2)

clean3 = clean2['StandingsLists']
# print(clean3)

#clean 3 had no keys and was a list. To access this portion set new variable to index 0
clean4 = clean3[0]

#Accessed further down and arrived at data desired
clean5 = clean4['DriverStandings']

#Used to capture Keys for dictionaries
#-------------------------------
# def getList(dict):
#     return dict.keys()

# dict = Constructors1
# print(getList(dict))
#-------------------------------

# Can I loop again within the two new dictionaries for Driver and Constructors if needed?

# for i in range(len(Driver)):
#     driverId = Driver[i]['driverId']
#     print(driverId)

# for j in range(len(clean5)):
#     Constructors = clean5[j]['Constructors']
#     # print(Constructors)

# for k  in range(len(clean5)): 
#     Position = clean5[k]['position']
#     # print(Position)

# for l in range(len(clean5)):
#     Points = clean5[l]['points']
#-------------------------------
# EXPLORED WITH XLSWRITER

# #create a workbook(xlsxwriter)

# workbook = xlsxwriter.Workbook('Formula1.xlsx')
# worksheet = workbook.add_worksheet()

# worksheet.write('A1', 'Hello world')
# workbook.save('f1.xlsx')
# workbook.close()
#-------------------------------

#Created a workbook

wb = Workbook()
ws = wb.active
#practice creating new worksheet
ws1 = wb.create_sheet("misc")
ws.title = "F1 Last Results Info"
#viewing worksheet names for workbook
print(wb.sheetnames)

#create memory of cells
for x in range(1,101):
    for y in range(1,101):
        ws.cell(row=x, column=y)

#Designate columns

ws['A1'] = "driverID"
# sheet['A1'].font = openpyxl.styles.Font(bold=True)

ws['B1'] = "code"
# sheet['B1'].font = openpyxl.styles.Font(bold=True)

ws['C1'] = "givenName"
# sheet['C1'].font = openpyxl.styles.Font(bold=True)

ws['D1'] = "familyName"
# sheet['D1'].font = openpyxl.styles.Font(bold=True)

ws['E1'] = "dateOfBirth"
# sheet['E1'].font = openpyxl.styles.Font(bold=True)

ws['F1'] = "Dnationality"
# sheet['F1'].font = openpyxl.styles.Font(bold=True)

ws['G1'] = "constructorId"
# sheet['G1'].font = openpyxl.styles.Font(bold=True)

ws['H1'] = "name"
# sheet['H1'].font = openpyxl.styles.Font(bold=True)

ws['I1'] = "Cnationality"
# sheet['I1'].font = openpyxl.styles.Font(bold=True)

ws['J1'] = "Position"
# sheet['I1'].font = openpyxl.styles.Font(bold=True)

ws['K1'] = "Points"
# sheet['I1'].font = openpyxl.styles.Font(bold=True)

for sheet in wb:
    print(sheet.title)

header_list = [ws['A1'], ws['B1'],ws['C1'], ws['D1'],ws['E1'], ws['F1'],ws['G1'], ws['H1'],ws['I1'],ws['J1'],ws['K1']]

# # # #Loop through the columns
for g in range(len(header_list)):
#     #change the font to bold
    header_list[g].font = openpyxl.styles.Font(bold=True)
    
# # #Loop through data from api
# #     #fill cell falues with appropriate data
counter = 2
# for each row in an excel spreadsheet
for i in range(len(clean5)):
    Driver = clean5[i]['Driver']
    del Driver['url']
    del Driver['permanentNumber']
    ws['A' + str(counter)] = Driver['driverId']
    ws['B' + str(counter)] = Driver['code']
    ws['C' + str(counter)] = Driver['givenName']
    ws['D' + str(counter)] = Driver['familyName']
    ws['E' + str(counter)] = Driver['dateOfBirth']
    ws['F' + str(counter)] = Driver['nationality']
    counter+=1

counter = 2
for j in range(len(clean5)):
    Constructors = clean5[j]['Constructors']
    Constructors1 = Constructors[0]
    del Constructors1['url']
    ws['G' + str(counter)] = Constructors1['constructorId']
    ws['H' + str(counter)] = Constructors1['name']
    ws['I' + str(counter)] = Constructors1['nationality']
    counter+=1
    # print(Constructors1)
    #type dict

counter = 2
for k  in range(len(clean5)): 
    Position = clean5[k]['position']
    ws['J' + str(counter)] = Position
    counter+=1
    # print(Position)
    #type str

counter = 2
for l in range(len(clean5)):
    Points = int(clean5[l]['points'])
    ws['K' + str(counter)] = Points
    counter+=1
    # print(Points)
    #type int

# # #Save the workbook
wb.save('Formula1.xlsx')

#check wb for information without having to load
for row in ws.iter_rows(max_row=20, max_col=11, values_only=True):
    print(row)

print(type(Driver))
print(type(Constructors1))
print(type(Position))
print(type(Points))